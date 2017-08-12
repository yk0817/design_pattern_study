# import sys
import openpyxl
import glob
import xlrd
import re
from abc import ABCMeta, abstractmethod

class ExcelTemplate(metaclass=ABCMeta):
    # デコレータをつけると、サブクラス抽象メソッドをオーバーライドしないとエラーになる。
    @abstractmethod
    def read_sheet():
        pass
    def get_filename(self,file):
        print(file + "の読み取りを行います。")

class XlsRead(ExcelTemplate):
    # xlsファイル読み込み
    def read_sheet(self,worksheet):
        # ワークシートを開く
        book = xlrd.open_workbook(worksheet)
        # シート名を取得 戻り値はリスト
        print(book.sheet_names())
        sheet_name = book.sheet_names()[0]
        # 一番目のシートを指定
        sheet_1 = book.sheet_by_name(sheet_name)
        print("値が入力されている 行列数：",sheet_1.ncols,sheet_1.nrows)
        print("1行目:1列目の値は:",sheet_1.cell(0, 0).value)



class MxlsRead(ExcelTemplate):
    # xlsxファイル読み込み
    def read_sheet(self,worksheet):
        book = openpyxl.load_workbook(worksheet)
        # シート名表示 戻り値はリスト
        print(book.get_sheet_names())
        sheet_name = book.get_sheet_names()[0]
        sheet_1 = book[sheet_name]
        print("値が入力されている 行列数:",sheet_1.max_column,sheet_1.max_row)
        print("2行目:1列目の値は:",sheet_1.cell(row=2,column=1).value)

def main():
    xls_instance = XlsRead()
    mxls_instance = MxlsRead()
    # 実行ファイルの一階層下にファイルが存在
    worksheets = glob.glob('../*')
    # 取得したエクセルリストに対して
    # xlsファイル、xlsmファイルに応じて切り替え
    for worksheet in worksheets:
        if re.match(r".+\.xls$",worksheet):
            xls_instance.get_filename(worksheet)
            xls_instance.read_sheet(worksheet)
            print("------------")
        elif re.match(r".+\.xlsx$",worksheet):
            mxls_instance.get_filename(worksheet)
            mxls_instance.read_sheet(worksheet)
            print("------------")
        else:
            pass

if __name__ == '__main__':
    main()
